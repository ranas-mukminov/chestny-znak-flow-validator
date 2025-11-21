from pathlib import Path

import io
import pytest
from openpyxl import Workbook

from cz_validator.io_import.xlsx_importer import load_wholesaler_export
from cz_validator.io_import.mapping_profiles import MappingProfile


def test_load_wholesaler_export(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["code", "gtin", "batch_id", "status", "owner_id"])
    ws.append(["WX1", "555", "B2", "IN_TRANSIT", "WHOLE"])
    ws.append(["WX2", "555", "B2", "AT_RETAIL", "WHOLE"])
    xlsx_path = tmp_path / "wholesaler.xlsx"
    wb.save(xlsx_path)

    profile = MappingProfile(
        column_mappings={
            "code": "code",
            "gtin": "gtin",
            "batch_id": "batch_id",
            "status": "status",
            "owner_id": "owner_id",
        }
    )
    codes = load_wholesaler_export(xlsx_path, profile)
    assert [c.code for c in codes] == ["WX1", "WX2"]
    assert codes[0].status.name == "IN_TRANSIT"


def test_load_wholesaler_export_closes_workbook(monkeypatch, tmp_path: Path):
    header = ["code", "gtin", "batch_id", "status", "owner_id"]
    data_row = ["WX1", "555", "B2", "IN_TRANSIT", "WHOLE"]
    closed = False

    class DummyFile(io.BytesIO):
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            self.close()
            return False

    class FakeCell:
        def __init__(self, value):
            self.value = value

    class FakeWorksheet:
        def iter_rows(self, min_row=1, max_row=None):
            if min_row == 1:
                yield [FakeCell(v) for v in header]
            else:
                yield [FakeCell(v) for v in data_row]

    class FakeWorkbook:
        def __init__(self):
            self.active = FakeWorksheet()

        def close(self):
            nonlocal closed
            closed = True

    monkeypatch.setattr(Path, "open", lambda _self, mode="rb": DummyFile(b""), raising=False)
    monkeypatch.setattr(
        "cz_validator.io_import.xlsx_importer.load_workbook", lambda _file: FakeWorkbook()
    )

    profile = MappingProfile(column_mappings={})
    load_wholesaler_export(tmp_path / "dummy.xlsx", profile)

    assert closed, "Workbook should be closed after processing"


def test_load_wholesaler_export_closes_workbook_on_error(monkeypatch, tmp_path: Path):
    closed = False

    class DummyFile(io.BytesIO):
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            self.close()
            return False

    class FakeCell:
        def __init__(self, value):
            self.value = value

    class FakeWorksheet:
        def iter_rows(self, min_row=1, max_row=None):
            if min_row == 1:
                yield [FakeCell("code")]
            else:
                raise RuntimeError("boom")

    class FakeWorkbook:
        def __init__(self):
            self.active = FakeWorksheet()

        def close(self):
            nonlocal closed
            closed = True

    monkeypatch.setattr(Path, "open", lambda _self, mode="rb": DummyFile(b""), raising=False)
    monkeypatch.setattr(
        "cz_validator.io_import.xlsx_importer.load_workbook", lambda _file: FakeWorkbook()
    )

    with pytest.raises(RuntimeError):
        load_wholesaler_export(tmp_path / "dummy.xlsx", MappingProfile())

    assert closed, "Workbook should be closed even when an error occurs"
