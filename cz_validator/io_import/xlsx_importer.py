from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook

from cz_validator.io_import.csv_importer import _infer_owner_type
from cz_validator.io_import.mapping_profiles import MappingProfile
from cz_validator.models import Code, Status


STATUS_MAP = {s.value: s for s in Status}


def load_wholesaler_export(path: Path, profile: MappingProfile) -> List[Code]:
    wb = load_workbook(path)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    codes: List[Code] = []
    for row in ws.iter_rows(min_row=2):
        data = {header[i]: (row[i].value if i < len(row) else None) for i in range(len(header))}
        code_val = data.get(profile.column_mappings.get("code", "code")) or data.get("code")
        if not code_val:
            continue
        status_val = data.get(profile.column_mappings.get("status", "status"))
        status = STATUS_MAP.get(status_val, Status.UNKNOWN)
        owner_raw = data.get(profile.column_mappings.get("owner_id", "owner_id"))
        codes.append(
            Code(
                code=code_val,
                gtin=data.get(profile.column_mappings.get("gtin", "gtin")),
                batch_id=data.get(profile.column_mappings.get("batch_id", "batch_id")),
                status=status,
                owner_type=_infer_owner_type(owner_raw),
                owner_id=owner_raw,
            )
        )
    return codes
